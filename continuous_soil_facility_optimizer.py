"""
Continuous Soil Remediation Facility Optimizer - Streamlit App
Determines optimal treatment cell configuration for continuous daily soil volumes
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import math

# ============================================================================
# Configuration and Constants
# ============================================================================

# Typical soil density: ~1.5 tons/CY (compacted)
# This can be adjusted based on soil type
SOIL_DENSITY_LB_PER_CY = 2700  # ~1.35 tons/CY

# ============================================================================
# Helper Functions
# ============================================================================

def calculate_workdays_in_cycle(start_date, duration_days, saturday_work, sunday_work):
    """Calculate actual working days considering weekends"""
    workdays = 0
    current = start_date
    
    for _ in range(duration_days):
        day_of_week = current.weekday()
        
        # Weekdays always count
        if day_of_week < 5:
            workdays += 1
        # Saturday
        elif day_of_week == 5 and saturday_work:
            workdays += 1
        # Sunday
        elif day_of_week == 6 and sunday_work:
            workdays += 1
        
        current += timedelta(days=1)
    
    return workdays


def calculate_calendar_days_for_workdays(target_workdays, start_date, saturday_work, sunday_work):
    """Calculate how many calendar days are needed to achieve target workdays"""
    workdays = 0
    calendar_days = 0
    current = start_date
    
    while workdays < target_workdays:
        day_of_week = current.weekday()
        
        # Weekdays always count
        if day_of_week < 5:
            workdays += 1
        # Saturday
        elif day_of_week == 5 and saturday_work:
            workdays += 1
        # Sunday
        elif day_of_week == 6 and sunday_work:
            workdays += 1
        
        calendar_days += 1
        current += timedelta(days=1)
    
    return calendar_days


def calculate_cell_dimensions(cell_volume_cy, depth_ft, aspect_ratio=2.0):
    """Calculate cell length and width from volume and depth
    Uses rectangular cells with consistent aspect ratio (Length:Width)
    Default aspect ratio is 2:1 (length is 2x width)
    """
    # Convert CY to cubic feet
    volume_cf = cell_volume_cy * 27
    
    # For a rectangular cell: Volume = L * W * D
    # With aspect ratio: L = aspect_ratio * W
    # Therefore: Volume = (aspect_ratio * W) * W * D = aspect_ratio * W^2 * D
    # Solving for W: W = sqrt(Volume / (aspect_ratio * D))
    width_ft = math.sqrt(volume_cf / (aspect_ratio * depth_ft))
    length_ft = aspect_ratio * width_ft
    
    return {
        'Length_ft': length_ft,
        'Width_ft': width_ft,
        'Depth_ft': depth_ft,
        'Volume_CY': cell_volume_cy,
        'Volume_CF': volume_cf,
        'Area_SF': length_ft * width_ft,
        'Aspect_Ratio': aspect_ratio
    }


def calculate_loading_time(cell_volume_cy, daily_load_capacity, saturday_work, sunday_work):
    """Calculate how long it takes to load a cell"""
    # Calculate workdays needed to load the cell
    workdays_needed = math.ceil(cell_volume_cy / daily_load_capacity)
    
    # Convert to calendar days considering weekend schedule
    calendar_days = calculate_calendar_days_for_workdays(
        workdays_needed,
        datetime.now(),  # Arbitrary start for calculation
        saturday_work,
        sunday_work
    )
    
    return {
        'workdays': workdays_needed,
        'calendar_days': calendar_days
    }


def calculate_total_cycle_time(cell_volume_cy, daily_incoming_volume, daily_equipment_capacity,
                               rip_days, treat_days, dry_days,
                               load_saturday, load_sunday,
                               rip_saturday, rip_sunday,
                               treat_saturday, treat_sunday,
                               dry_saturday, dry_sunday,
                               unload_saturday, unload_sunday):
    """Calculate total time for a complete cell cycle
    
    Key: Loading time is based on INCOMING VOLUME, not equipment capacity.
    You can only load what actually arrives each day.
    Unloading time is based on equipment capacity.
    """
    
    # Loading time - constrained by incoming volume, not equipment
    load_workdays = math.ceil(cell_volume_cy / daily_incoming_volume)
    load_calendar_days = calculate_calendar_days_for_workdays(
        load_workdays, datetime.now(), load_saturday, load_sunday
    )
    
    # Rip time
    rip_calendar_days = calculate_calendar_days_for_workdays(
        rip_days, datetime.now(), rip_saturday, rip_sunday
    )
    
    # Treatment time
    treat_calendar_days = calculate_calendar_days_for_workdays(
        treat_days, datetime.now(), treat_saturday, treat_sunday
    )
    
    # Drying time
    dry_calendar_days = calculate_calendar_days_for_workdays(
        dry_days, datetime.now(), dry_saturday, dry_sunday
    )
    
    # Unloading time - constrained by equipment capacity
    unload_workdays = math.ceil(cell_volume_cy / daily_equipment_capacity)
    unload_calendar_days = calculate_calendar_days_for_workdays(
        unload_workdays, datetime.now(), unload_saturday, unload_sunday
    )
    
    total_calendar_days = (load_calendar_days + rip_calendar_days + 
                          treat_calendar_days + dry_calendar_days + 
                          unload_calendar_days)
    
    return {
        'load_calendar_days': load_calendar_days,
        'load_workdays': load_workdays,
        'rip_calendar_days': rip_calendar_days,
        'treat_calendar_days': treat_calendar_days,
        'dry_calendar_days': dry_calendar_days,
        'unload_calendar_days': unload_calendar_days,
        'unload_workdays': unload_workdays,
        'total_calendar_days': total_calendar_days
    }


def calculate_cells_needed(daily_volume, cell_volume, cycle_days, buffer_factor=1.1):
    """Calculate minimum number of cells needed for continuous operation
    
    Args:
        daily_volume: CY/day arriving at facility
        cell_volume: CY per cell
        cycle_days: Total calendar days for complete cell cycle
        buffer_factor: Safety factor (default 1.1 = 10% buffer)
    """
    # Volume accumulation during one complete cycle
    volume_per_cycle = daily_volume * cycle_days
    
    # Minimum cells needed (theoretical)
    min_cells_theoretical = volume_per_cycle / cell_volume
    
    # Add buffer and round up
    min_cells_with_buffer = math.ceil(min_cells_theoretical * buffer_factor)
    
    return {
        'min_cells_theoretical': min_cells_theoretical,
        'min_cells_with_buffer': min_cells_with_buffer,
        'volume_per_cycle': volume_per_cycle
    }


def find_max_daily_volume(num_cells, cell_volume, daily_equipment_capacity,
                          phase_params, weekend_params, simulation_days=120):
    """Binary search to find the maximum daily volume a configuration can handle
    while maintaining continuous operation (zero idle days AND zero queue buildup).
    
    Returns the maximum CY/day this config can process sustainably.
    """
    
    low, high = 50, 1500  # Search range for daily volume
    max_found = 0
    
    while low <= high:
        mid = (low + high) // 2
        idle_days, max_waiting = simulate_for_idle_days(
            num_cells, cell_volume, mid, daily_equipment_capacity,
            phase_params, weekend_params, simulation_days
        )
        
        # Must have BOTH zero idle days AND zero queue buildup to be sustainable
        if idle_days == 0 and max_waiting == 0:
            max_found = mid
            low = mid + 1
        else:
            high = mid - 1
    
    return max_found


def optimize_cell_configuration(daily_volume_cy, daily_equipment_capacity,
                                phase_params, weekend_params,
                                min_cell_volume=100, max_cell_volume=5000, 
                                step_size=100):
    """Find optimal cell configurations prioritized by:
    1. Zero idle days (never turn away work)
    2. Fewest cells (minimize capital cost)
    3. Smallest cell size (if tied on cells)
    
    Runs simulation to determine actual idle days for each configuration.
    Only returns configurations that can handle the planned daily volume.
    Limited to top 10 results.
    """
    
    results = []
    
    # Test different cell sizes
    for cell_volume in range(min_cell_volume, max_cell_volume + 1, step_size):
        
        # Calculate cycle time
        cycle_info = calculate_total_cycle_time(
            cell_volume,
            daily_volume_cy,  # Loading constrained by incoming volume
            daily_equipment_capacity,
            phase_params['rip_days'],
            phase_params['treat_days'],
            phase_params['dry_days'],
            weekend_params['load_saturday'],
            weekend_params['load_sunday'],
            weekend_params['rip_saturday'],
            weekend_params['rip_sunday'],
            weekend_params['treat_saturday'],
            weekend_params['treat_sunday'],
            weekend_params['dry_saturday'],
            weekend_params['dry_sunday'],
            weekend_params['unload_saturday'],
            weekend_params['unload_sunday']
        )
        
        # Test different numbers of cells (2 to 12)
        for num_cells in range(2, 13):
            
            # Run simulation at the PLANNED volume to check if it's sustainable
            idle_days, max_waiting = simulate_for_idle_days(
                num_cells, cell_volume, daily_volume_cy, daily_equipment_capacity,
                phase_params, weekend_params, simulation_days=180  # Longer sim for accuracy
            )
            
            # Skip configurations that can't sustain the planned volume
            # Must have zero idle days AND zero queue buildup
            if idle_days > 0 or max_waiting > 0:
                continue
            
            # Calculate maximum daily volume this config can handle (for display)
            max_daily_volume = find_max_daily_volume(
                num_cells, cell_volume, daily_equipment_capacity,
                phase_params, weekend_params
            )
            
            # If binary search max is less than planned (due to non-monotonic behavior),
            # use planned volume since we already verified it works
            if max_daily_volume < daily_volume_cy:
                max_daily_volume = daily_volume_cy
            
            # Calculate buffer/headroom
            buffer_cy = max_daily_volume - daily_volume_cy
            buffer_pct = (buffer_cy / daily_volume_cy) * 100
            
            # Calculate metrics
            total_capacity = cell_volume * num_cells
            
            # Score: Primary=idle_days, Secondary=num_cells, Tertiary=cell_volume
            # Lower is better for all three
            # Weight idle_days very heavily - it's the primary constraint
            score = (
                idle_days * 100000 +  # Primary: zero idle days is critical
                num_cells * 1000 +     # Secondary: fewer cells better
                cell_volume            # Tertiary: smaller cells if tied
            )
            
            results.append({
                'cell_volume_cy': cell_volume,
                'num_cells': num_cells,
                'load_days': cycle_info['load_calendar_days'],
                'cycle_days': cycle_info['total_calendar_days'],
                'total_capacity_cy': total_capacity,
                'idle_days': idle_days,
                'max_waiting_cy': max_waiting,
                'max_daily_volume': max_daily_volume,
                'buffer_cy': buffer_cy,
                'buffer_pct': buffer_pct,
                'score': score,
                'cycle_breakdown': cycle_info
            })
            
            # If we found zero idle days, no need to try more cells for this size
            if idle_days == 0:
                break
    
    if not results:
        return None
    
    # Sort by score (lower is better) and limit to top 10
    results_df = pd.DataFrame(results)
    results_df = results_df.sort_values('score').reset_index(drop=True)
    results_df = results_df.head(10)
    
    return results_df


def simulate_for_idle_days(num_cells, cell_volume, daily_volume_cy, daily_equipment_capacity,
                           phase_params, weekend_params, simulation_days=90):
    """Quick simulation to count idle days (days where soil arrives but can't be loaded)
    
    Key rules:
    - Loading rate = incoming volume (e.g., 300 CY/day) - MUST keep up with trucks
    - Unloading rate = remaining equipment capacity (e.g., 750 - 300 = 450 CY/day)
    - Both happen on work days when there's work to do
    - Equipment capacity is the ceiling on total dirt moved per day
    """
    
    def is_work_day(date, phase_type):
        day_of_week = date.weekday()
        is_saturday = day_of_week == 5
        is_sunday = day_of_week == 6
        
        if phase_type == 'load':
            if is_saturday and not weekend_params['load_saturday']:
                return False
            if is_sunday and not weekend_params['load_sunday']:
                return False
        elif phase_type == 'unload':
            if is_saturday and not weekend_params['unload_saturday']:
                return False
            if is_sunday and not weekend_params['unload_sunday']:
                return False
        elif phase_type == 'rip':
            if is_saturday and not weekend_params['rip_saturday']:
                return False
            if is_sunday and not weekend_params['rip_sunday']:
                return False
        elif phase_type == 'treat':
            if is_saturday and not weekend_params['treat_saturday']:
                return False
            if is_sunday and not weekend_params['treat_sunday']:
                return False
        elif phase_type == 'dry':
            if is_saturday and not weekend_params['dry_saturday']:
                return False
            if is_sunday and not weekend_params['dry_sunday']:
                return False
        return True
    
    class CellState:
        def __init__(self, cell_num):
            self.cell_num = cell_num
            self.phase = 'Empty'
            self.soil_volume = 0
            self.phase_workdays_completed = 0
            self.pending_transition = None
            self.flip_num = 0
    
    cells = [CellState(i) for i in range(num_cells)]
    current_date = datetime(2025, 12, 1)
    
    active_loading_cell = None
    active_unloading_cell = None
    soil_waiting = 0
    max_waiting = 0
    idle_days = 0
    next_flip_num = 1
    
    # Calculate daily rates
    daily_load_rate = daily_volume_cy  # Loading matches incoming
    daily_unload_rate = daily_equipment_capacity - daily_volume_cy  # Unloading gets the rest
    
    for day in range(simulation_days):
        # Process pending transitions
        for i, cell in enumerate(cells):
            if cell.pending_transition:
                cell.phase = cell.pending_transition
                cell.pending_transition = None
                cell.phase_workdays_completed = 0
                if cell.phase == 'Rip' and active_loading_cell == i:
                    active_loading_cell = None
                if cell.phase == 'Empty' and active_unloading_cell == i:
                    active_unloading_cell = None
                    cell.flip_num = 0
        
        # Soil arrives on loading work days
        if is_work_day(current_date, 'load'):
            soil_waiting += daily_volume_cy
        
        loaded_today = False
        can_load = is_work_day(current_date, 'load')
        can_unload = is_work_day(current_date, 'unload')
        
        # ============================================================
        # UNLOADING FIRST - so cells become available for loading
        # ============================================================
        if can_unload:
            # Find or assign active unloading cell (lowest flip number ready)
            if active_unloading_cell is None:
                ready_cells = [(i, cells[i]) for i in range(num_cells) 
                              if cells[i].phase == 'ReadyToUnload']
                if ready_cells:
                    ready_cells.sort(key=lambda x: x[1].flip_num)
                    active_unloading_cell = ready_cells[0][0]
                    cells[active_unloading_cell].phase = 'Unloading'
            
            # Unload at surplus rate
            if active_unloading_cell is not None:
                cell = cells[active_unloading_cell]
                if cell.phase == 'Unloading':
                    unload_amount = min(cell.soil_volume, daily_unload_rate)
                    if unload_amount > 0:
                        cell.soil_volume -= unload_amount
                    if cell.soil_volume <= 0:
                        # Cell is now IMMEDIATELY empty and available for loading today
                        cell.phase = 'Empty'
                        cell.flip_num = 0
                        active_unloading_cell = None
        
        # ============================================================
        # LOADING - now any just-emptied cells are available
        # ============================================================
        if can_load:
            remaining_load_capacity = daily_load_rate
            
            while remaining_load_capacity > 0 and soil_waiting > 0:
                # Find or assign active loading cell
                if active_loading_cell is None:
                    for i in range(num_cells):
                        if cells[i].phase == 'Empty':
                            active_loading_cell = i
                            cells[i].phase = 'Loading'
                            cells[i].soil_volume = 0
                            cells[i].flip_num = next_flip_num
                            next_flip_num += 1
                            break
                
                # If no empty cell available, stop
                if active_loading_cell is None:
                    break
                
                # Load the active cell
                cell = cells[active_loading_cell]
                if cell.phase == 'Loading':
                    space_remaining = cell_volume - cell.soil_volume
                    load_amount = min(space_remaining, remaining_load_capacity, soil_waiting)
                    if load_amount > 0:
                        cell.soil_volume += load_amount
                        soil_waiting -= load_amount
                        remaining_load_capacity -= load_amount
                        loaded_today = True
                    
                    # Check if cell is full - transition and clear active
                    if cell.soil_volume >= cell_volume:
                        cell.pending_transition = 'Rip'
                        active_loading_cell = None
            
            # Count as idle day if soil arrived but we couldn't load any
            if not loaded_today:
                idle_days += 1
        
        max_waiting = max(max_waiting, soil_waiting)
        
        # Treatment phases (don't use equipment)
        for cell in cells:
            if cell.phase == 'Rip':
                if is_work_day(current_date, 'rip'):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['rip_days']:
                        cell.pending_transition = 'Treat'
            elif cell.phase == 'Treat':
                if is_work_day(current_date, 'treat'):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['treat_days']:
                        cell.pending_transition = 'Dry'
            elif cell.phase == 'Dry':
                if is_work_day(current_date, 'dry'):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['dry_days']:
                        cell.pending_transition = 'ReadyToUnload'
        
        current_date += timedelta(days=1)
    
    return idle_days, max_waiting


def is_valid_work_day(date, phase_name, weekend_params):
    """Check if a date is valid for work based on weekend rules"""
    day_of_week = date.weekday()  # Monday=0, Sunday=6
    
    # Weekdays (Mon-Fri) are always valid
    if day_of_week < 5:
        return True
    
    # Saturday
    if day_of_week == 5:
        return weekend_params.get(f'{phase_name.lower()}_saturday', False)
    
    # Sunday
    if day_of_week == 6:
        return weekend_params.get(f'{phase_name.lower()}_sunday', False)
    
    return False


def simulate_facility_schedule(config, daily_volume_cy, daily_equipment_capacity, 
                               phase_params, weekend_params,
                               start_date, simulation_days=180):
    """Simulate facility operations
    
    Key Rules:
    - Loading rate = incoming volume (e.g., 300 CY/day) - MUST keep up with trucks
    - Unloading rate = remaining equipment capacity (e.g., 750 - 300 = 450 CY/day)
    - Both happen on work days when there's work to do
    - Equipment capacity (750) is the ceiling on total dirt moved per day
    - Phase transitions happen at the START of the next day
    """
    
    cell_volume = config['cell_volume_cy']
    num_cells = int(config['num_cells'])
    
    # Calculate daily rates
    daily_load_rate = daily_volume_cy  # Loading matches incoming
    daily_unload_rate = daily_equipment_capacity - daily_volume_cy  # Unloading gets the rest
    
    # Initialize cell tracking
    class CellState:
        def __init__(self, cell_num):
            self.cell_num = cell_num
            self.phase = 'Empty'
            self.soil_volume = 0
            self.phase_workdays_completed = 0
            self.flip_num = 0
            self.pending_transition = None
            
    cells = {i: CellState(i) for i in range(1, num_cells + 1)}
    
    # Track daily activities
    schedule = []
    current_date = start_date
    
    # Track which cell is actively loading/unloading (only one at a time!)
    active_loading_cell = None
    active_unloading_cell = None
    
    # Soil waiting to be loaded (accumulates if can't load fast enough)
    soil_waiting = 0
    
    # Tracking totals
    total_soil_loaded = 0
    total_soil_unloaded = 0
    next_flip_num = 1
    
    for day in range(simulation_days):
        
        # ============================================================
        # START OF DAY: Process pending transitions from yesterday
        # ============================================================
        for cell_num in range(1, num_cells + 1):
            cell = cells[cell_num]
            if cell.pending_transition:
                cell.phase = cell.pending_transition
                cell.pending_transition = None
                cell.phase_workdays_completed = 0
                
                # Clear active references if cell transitioned out
                if cell.phase == 'Rip' and active_loading_cell == cell_num:
                    active_loading_cell = None
                if cell.phase == 'Empty' and active_unloading_cell == cell_num:
                    active_unloading_cell = None
        
        # ============================================================
        # DAILY INCOMING: Add today's incoming soil to waiting pile
        # ============================================================
        if is_valid_work_day(current_date, 'load', weekend_params):
            soil_waiting += daily_volume_cy
        
        day_record = {
            'Date': current_date,
            'DayName': current_date.strftime('%A'),
        }
        
        daily_soil_in = 0
        daily_soil_out = 0
        
        can_load = is_valid_work_day(current_date, 'load', weekend_params)
        can_unload = is_valid_work_day(current_date, 'unload', weekend_params)
        
        # ============================================================
        # UNLOADING FIRST - so cells become available for loading
        # ============================================================
        cell_unload_amounts = {}  # Track which cell unloaded and how much
        
        if can_unload:
            # Find or assign active unloading cell (lowest flip number ready)
            if active_unloading_cell is None:
                ready_cells = [(cell_num, cells[cell_num]) for cell_num in range(1, num_cells + 1) 
                              if cells[cell_num].phase == 'ReadyToUnload']
                if ready_cells:
                    ready_cells.sort(key=lambda x: x[1].flip_num)
                    active_unloading_cell = ready_cells[0][0]
                    cells[active_unloading_cell].phase = 'Unloading'
            
            # Unload at surplus rate
            if active_unloading_cell is not None:
                cell = cells[active_unloading_cell]
                if cell.phase == 'Unloading':
                    unload_amount = min(cell.soil_volume, daily_unload_rate)
                    if unload_amount > 0:
                        cell.soil_volume -= unload_amount
                        daily_soil_out = unload_amount
                        total_soil_unloaded += unload_amount
                        cell_unload_amounts[active_unloading_cell] = unload_amount
                    if cell.soil_volume <= 0:
                        # Cell is now IMMEDIATELY empty and available for loading today
                        cell.phase = 'Empty'
                        cell.flip_num = 0
                        active_unloading_cell = None
        
        # ============================================================
        # LOADING - now any just-emptied cells are available
        # ============================================================
        cell_load_amounts = {}  # Track how much each cell loaded today
        
        if can_load:
            remaining_load_capacity = daily_load_rate  # Can load up to incoming rate
            
            while remaining_load_capacity > 0 and soil_waiting > 0:
                # Find or assign active loading cell
                if active_loading_cell is None:
                    for cell_num in range(1, num_cells + 1):
                        cell = cells[cell_num]
                        if cell.phase == 'Empty':
                            cell.phase = 'Loading'
                            cell.flip_num = next_flip_num
                            next_flip_num += 1
                            cell.soil_volume = 0
                            active_loading_cell = cell_num
                            break
                
                # If no empty cell available, stop
                if active_loading_cell is None:
                    break
                
                # Load the active cell
                cell = cells[active_loading_cell]
                if cell.phase == 'Loading':
                    space_remaining = cell_volume - cell.soil_volume
                    load_amount = min(space_remaining, remaining_load_capacity, soil_waiting)
                    if load_amount > 0:
                        cell.soil_volume += load_amount
                        soil_waiting -= load_amount
                        remaining_load_capacity -= load_amount
                        daily_soil_in += load_amount
                        total_soil_loaded += load_amount
                        # Track per-cell loading
                        cell_load_amounts[active_loading_cell] = cell_load_amounts.get(active_loading_cell, 0) + load_amount
                    
                    # Check if cell is full - transition and clear active so next iteration finds new cell
                    if cell.soil_volume >= cell_volume:
                        cell.pending_transition = 'Rip'
                        active_loading_cell = None  # Clear so we find next empty cell
        
        # ============================================================
        # Progress treatment phases for ALL cells
        # ============================================================
        for cell_num in range(1, num_cells + 1):
            cell = cells[cell_num]
            
            if cell.phase == 'Rip':
                if is_valid_work_day(current_date, 'rip', weekend_params):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['rip_days']:
                        cell.pending_transition = 'Treat'
                        
            elif cell.phase == 'Treat':
                if is_valid_work_day(current_date, 'treat', weekend_params):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['treat_days']:
                        cell.pending_transition = 'Dry'
                        
            elif cell.phase == 'Dry':
                if is_valid_work_day(current_date, 'dry', weekend_params):
                    cell.phase_workdays_completed += 1
                    if cell.phase_workdays_completed >= phase_params['dry_days']:
                        cell.pending_transition = 'ReadyToUnload'
        
        # ============================================================
        # Record cell states for this day
        # ============================================================
        for cell_num in range(1, num_cells + 1):
            cell = cells[cell_num]
            
            # Format phase display
            # If cell both unloaded AND loaded today, show Load (it's the ending activity)
            if cell_num in cell_load_amounts:
                phase_display = f"Load ({int(cell_load_amounts[cell_num])})"
            # Check if this cell unloaded today (even if it's now Empty)
            elif cell_num in cell_unload_amounts:
                phase_display = f"Unload ({int(cell_unload_amounts[cell_num])})"
            elif cell.phase == 'Loading':
                phase_display = ''  # Loading but no soil loaded today
            elif cell.phase == 'Unloading':
                phase_display = ''  # Unloading but nothing unloaded today
            elif cell.phase == 'Rip':
                phase_display = 'Rip'
            elif cell.phase == 'Treat':
                phase_display = 'Treat'
            elif cell.phase == 'Dry':
                phase_display = 'Dry'
            elif cell.phase == 'ReadyToUnload':
                phase_display = ''
            elif cell.phase == 'Empty':
                phase_display = ''
            else:
                phase_display = cell.phase
            
            day_record[f'Cell_{cell_num}_Phase'] = phase_display
        
        day_record['SoilIn'] = daily_soil_in
        day_record['SoilOut'] = daily_soil_out
        day_record['SoilWaiting'] = soil_waiting
        day_record['CumSoilIn'] = total_soil_loaded
        day_record['CumSoilOut'] = total_soil_unloaded
        
        schedule.append(day_record)
        current_date += timedelta(days=1)
    
    return pd.DataFrame(schedule)


# ============================================================================
# Main Application
# ============================================================================

def main():
    st.set_page_config(
        page_title="Continuous Soil Facility Optimizer",
        page_icon="🏭",
        layout="wide"
    )
    
    st.title("🏭 Continuous Soil Remediation Facility Optimizer")
    st.markdown("""
    This tool calculates the optimal treatment cell configuration for a continuous-flow 
    soil remediation facility based on daily incoming volume.
    """)
    
    # Initialize session state
    if 'optimization_run' not in st.session_state:
        st.session_state.optimization_run = False
    if 'results_df' not in st.session_state:
        st.session_state.results_df = None
    if 'selected_config_index' not in st.session_state:
        st.session_state.selected_config_index = None
    if 'parameters' not in st.session_state:
        st.session_state.parameters = {}
    if 'schedule_generated' not in st.session_state:
        st.session_state.schedule_generated = False
    if 'schedule_df' not in st.session_state:
        st.session_state.schedule_df = None
    if 'schedule_start' not in st.session_state:
        st.session_state.schedule_start = None
    
    # Sidebar for inputs
    with st.sidebar:
        st.header("⚙️ Facility Parameters")
        
        # Daily Volume
        st.subheader("Incoming Volume")
        daily_volume = st.number_input(
            "Daily Soil Volume (CY/day)",
            min_value=10,
            max_value=5000,
            value=300,
            step=10,
            help="Average daily volume of soil arriving at facility"
        )
        
        # Equipment Capacity
        st.subheader("Equipment Capacity")
        daily_equipment_capacity = st.number_input(
            "Daily Equipment Capacity (CY/day)",
            min_value=50,
            max_value=2000,
            value=750,
            step=25,
            help="Total soil that can be moved per day (loading + unloading combined). Equipment is shared - prioritizes lowest-numbered active cell."
        )
        
        # Phase Durations
        st.subheader("Treatment Phase Durations (Work Days)")
        rip_days = st.number_input("Rip Duration (days)", min_value=1, value=1, step=1)
        treat_days = st.number_input("Treatment Duration (days)", min_value=1, value=3, step=1)
        dry_days = st.number_input("Drying Duration (days)", min_value=1, value=5, step=1)
        
        # Weekend Working
        st.subheader("Weekend Operations")
        
        st.markdown("**Loading**")
        load_saturday = st.checkbox("Work Saturdays (Load)", value=False)
        load_sunday = st.checkbox("Work Sundays (Load)", value=False)
        
        st.markdown("**Rip**")
        rip_saturday = st.checkbox("Work Saturdays (Rip)", value=True)
        rip_sunday = st.checkbox("Work Sundays (Rip)", value=True)
        
        st.markdown("**Treatment**")
        treat_saturday = st.checkbox("Work Saturdays (Treat)", value=True, 
                                    help="Treatment often continues 7 days/week")
        treat_sunday = st.checkbox("Work Sundays (Treat)", value=True,
                                   help="Treatment often continues 7 days/week")
        
        st.markdown("**Drying**")
        dry_saturday = st.checkbox("Work Saturdays (Dry)", value=True,
                                  help="Drying is passive, counts all days")
        dry_sunday = st.checkbox("Work Sundays (Dry)", value=True,
                                help="Drying is passive, counts all days")
        
        st.markdown("**Unloading**")
        unload_saturday = st.checkbox("Work Saturdays (Unload)", value=False)
        unload_sunday = st.checkbox("Work Sundays (Unload)", value=False)
        
        # Optimization Parameters
        st.subheader("Optimization Constraints")
        
        min_cell_size = st.number_input(
            "Min Cell Size (CY)",
            min_value=100,
            max_value=5000,
            value=900,
            step=100,
            help="Minimum cell size to consider"
        )
        
        max_cell_size = st.number_input(
            "Max Cell Size (CY)",
            min_value=500,
            max_value=10000,
            value=5000,
            step=100,
            help="Maximum cell size to consider"
        )
        
        run_button = st.button("🔍 Optimize Configuration", type="primary", use_container_width=True)
    
    # Main content area
    if run_button:
        with st.spinner("Optimizing cell configuration..."):
            
            # Prepare parameters
            phase_params = {
                'rip_days': rip_days,
                'treat_days': treat_days,
                'dry_days': dry_days
            }
            
            weekend_params = {
                'load_saturday': load_saturday,
                'load_sunday': load_sunday,
                'rip_saturday': rip_saturday,
                'rip_sunday': rip_sunday,
                'treat_saturday': treat_saturday,
                'treat_sunday': treat_sunday,
                'dry_saturday': dry_saturday,
                'dry_sunday': dry_sunday,
                'unload_saturday': unload_saturday,
                'unload_sunday': unload_sunday
            }
            
            # Run optimization
            results_df = optimize_cell_configuration(
                daily_volume,
                daily_equipment_capacity,
                phase_params,
                weekend_params,
                min_cell_volume=min_cell_size,
                max_cell_volume=max_cell_size,
                step_size=100
            )
            
            # Store in session state
            st.session_state.results_df = results_df
            st.session_state.selected_config_index = 0 if results_df is not None and len(results_df) > 0 else None
            st.session_state.optimization_run = True
            # Clear any previous schedule when new optimization runs
            st.session_state.schedule_generated = False
            st.session_state.schedule_df = None
            st.session_state.schedule_start = None
            st.session_state.parameters = {
                'daily_volume': daily_volume,
                'daily_equipment_capacity': daily_equipment_capacity,
                'phase_params': phase_params,
                'weekend_params': weekend_params,
                'rip_days': rip_days,
                'treat_days': treat_days,
                'dry_days': dry_days
            }
    
    # Display results if optimization has been run
    if st.session_state.optimization_run and st.session_state.results_df is not None:
        results_df = st.session_state.results_df
        
        # Retrieve parameters
        daily_volume = st.session_state.parameters['daily_volume']
        daily_equipment_capacity = st.session_state.parameters['daily_equipment_capacity']
        phase_params = st.session_state.parameters['phase_params']
        weekend_params = st.session_state.parameters['weekend_params']
        rip_days = st.session_state.parameters['rip_days']
        treat_days = st.session_state.parameters['treat_days']
        dry_days = st.session_state.parameters['dry_days']
        
        if results_df is None or len(results_df) == 0:
            st.error("❌ No valid configurations found. Try adjusting constraints.")
            st.info("Suggestions: Increase max cell size, increase equipment capacity, or increase max loading days")
        else:
            # Count zero-idle configurations
            zero_idle_configs = results_df[results_df['idle_days'] == 0]
            
            if len(zero_idle_configs) > 0:
                st.success(f"✅ Found {len(zero_idle_configs)} configurations with continuous operation (zero idle days)")
            else:
                st.warning(f"⚠️ No configurations found with zero idle days. Showing {len(results_df)} configurations ranked by idle days.")
            
            # Summary stats
            st.subheader("📊 Configuration Summary")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if len(zero_idle_configs) > 0:
                    best = zero_idle_configs.iloc[0]
                    st.metric(
                        "🏆 Best Configuration",
                        f"{int(best['num_cells'])} × {int(best['cell_volume_cy'])} CY",
                        help="Fewest cells with zero idle days"
                    )
                else:
                    best = results_df.iloc[0]
                    st.metric(
                        "Best Available",
                        f"{int(best['num_cells'])} × {int(best['cell_volume_cy'])} CY",
                        f"{int(best['idle_days'])} idle days"
                    )
            
            with col2:
                st.metric(
                    "Cell Range",
                    f"{results_df['num_cells'].min():.0f} - {results_df['num_cells'].max():.0f} cells",
                    help="Range of cell counts analyzed"
                )
            
            with col3:
                st.metric(
                    "Size Range",
                    f"{results_df['cell_volume_cy'].min():.0f} - {results_df['cell_volume_cy'].max():.0f} CY",
                    help="Range of cell volumes analyzed"
                )
            
            with col4:
                if len(zero_idle_configs) > 0:
                    st.metric(
                        "Zero-Idle Options",
                        f"{len(zero_idle_configs)} configs",
                        help="Configurations that never turn away work"
                    )
                else:
                    st.metric(
                        "Min Idle Days",
                        f"{results_df['idle_days'].min():.0f} days",
                        help="Fewest idle days found"
                    )
            
            # Configuration Selection (simplified)
            st.subheader("📋 Select Configuration")
            
            col1, col2 = st.columns([3, 1])
            
            with col1:
                selected_index = st.selectbox(
                    "Choose configuration:",
                    options=range(len(results_df)),
                    format_func=lambda x: f"{int(results_df.iloc[x]['num_cells'])} × {int(results_df.iloc[x]['cell_volume_cy'])} CY (max {int(results_df.iloc[x]['max_daily_volume'])} CY/day)",
                    index=0,
                    key="config_selector"
                )
                st.session_state.selected_config_index = selected_index
            
            with col2:
                idle = results_df.iloc[selected_index]['idle_days']
                if idle == 0:
                    st.success("✅ Continuous")
                else:
                    st.warning(f"⚠️ {int(idle)} Idle")
            
            # Display selected configuration details
            selected_config = results_df.iloc[selected_index]
            
            detail_col1, detail_col2, detail_col3, detail_col4 = st.columns(4)
            
            with detail_col1:
                st.metric("Cell Size", f"{selected_config['cell_volume_cy']:.0f} CY")
                st.metric("Number of Cells", f"{int(selected_config['num_cells'])}")
            
            with detail_col2:
                st.metric("Total Capacity", f"{selected_config['total_capacity_cy']:.0f} CY")
                st.metric("Full Cycle", f"{selected_config['cycle_days']:.0f} days")
            
            with detail_col3:
                st.metric("Max Daily Volume", f"{int(selected_config['max_daily_volume'])} CY/day")
                buffer_pct = selected_config['buffer_pct']
                st.metric("Buffer Capacity", f"+{buffer_pct:.0f}%")
            
            with detail_col4:
                st.metric("Idle Days", f"{int(selected_config['idle_days'])}")
                if selected_config['idle_days'] == 0:
                    st.success("✅ Continuous Operation")
                else:
                    st.warning(f"⚠️ Will turn away work")
            
            # Auto-generate and display schedule
            st.markdown("---")
            st.subheader("📅 Treatment Schedule")
            
            col1, col2 = st.columns([2, 1])
            with col1:
                schedule_start_date = st.date_input(
                    "Schedule Start Date",
                    value=datetime.now().date(),
                    help="First day of facility operations",
                    key="schedule_start_date_input"
                )
            with col2:
                schedule_days = st.number_input(
                    "Days to Schedule",
                    min_value=30,
                    max_value=365,
                    value=90,
                    step=30,
                    help="Number of days to simulate",
                    key="schedule_days_number"
                )
            
            # Generate schedule
            schedule_start = datetime.combine(schedule_start_date, datetime.min.time())
            
            schedule_df = simulate_facility_schedule(
                selected_config,
                daily_volume,
                daily_equipment_capacity,
                phase_params,
                weekend_params,
                schedule_start,
                int(schedule_days)
            )
            
            # Store in session state for export
            st.session_state.schedule_df = schedule_df
            st.session_state.schedule_generated = True
            st.session_state.schedule_start = schedule_start
            
            # Create display columns
            display_cols = ['Date', 'DayName']
            for i in range(1, int(selected_config['num_cells']) + 1):
                display_cols.append(f'Cell_{i}_Phase')
            display_cols.extend(['SoilIn', 'SoilOut', 'SoilWaiting'])
            
            # Prepare display dataframe
            display_schedule = schedule_df[display_cols].copy()
            display_schedule['Date'] = display_schedule['Date'].dt.strftime('%m/%d/%Y')
            
            # Rename columns for cleaner display
            rename_map = {'DayName': 'Day', 'SoilIn': 'In', 'SoilOut': 'Out', 'SoilWaiting': 'Wait'}
            for i in range(1, int(selected_config['num_cells']) + 1):
                rename_map[f'Cell_{i}_Phase'] = f'Cell {i}'
            display_schedule = display_schedule.rename(columns=rename_map)
            
            # Define phase colors (matching Excel export)
            phase_colors = {
                'Load': '#8ED973',    # Green
                'Rip': '#83CCEB',     # Light blue
                'Treat': '#FFC000',   # Gold/Orange
                'Dry': '#F2CEEF',     # Pink
                'Unload': '#00B0F0',  # Bright blue
            }
            
            def style_schedule(df):
                """Apply colors to schedule dataframe"""
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                
                # Get cell columns
                cell_cols = [col for col in df.columns if col.startswith('Cell')]
                
                for idx in df.index:
                    # Check if Sunday for row highlighting
                    is_sunday = df.loc[idx, 'Day'] == 'Sunday'
                    
                    for col in df.columns:
                        if col in cell_cols:
                            # Color based on phase
                            val = str(df.loc[idx, col]) if pd.notna(df.loc[idx, col]) else ''
                            for phase, color in phase_colors.items():
                                if phase in val:
                                    styles.loc[idx, col] = f'background-color: {color}; text-align: center'
                                    break
                            else:
                                if is_sunday:
                                    styles.loc[idx, col] = 'background-color: #FFFF00; text-align: center'
                                else:
                                    styles.loc[idx, col] = 'text-align: center'
                        elif is_sunday and col not in ['Date']:
                            styles.loc[idx, col] = 'background-color: #FFFF00'
                
                return styles
            
            # Apply styling
            styled_schedule = display_schedule.style.apply(lambda x: style_schedule(display_schedule), axis=None)
            
            # Display the schedule
            st.dataframe(
                styled_schedule,
                use_container_width=True,
                hide_index=True,
                height=600
            )
            
            # Legend
            st.markdown("**Legend:**")
            legend_cols = st.columns(5)
            with legend_cols[0]:
                st.markdown('<div style="background-color: #8ED973; padding: 5px; text-align: center; border-radius: 3px;">Load</div>', unsafe_allow_html=True)
            with legend_cols[1]:
                st.markdown('<div style="background-color: #83CCEB; padding: 5px; text-align: center; border-radius: 3px;">Rip</div>', unsafe_allow_html=True)
            with legend_cols[2]:
                st.markdown('<div style="background-color: #FFC000; padding: 5px; text-align: center; border-radius: 3px;">Treat</div>', unsafe_allow_html=True)
            with legend_cols[3]:
                st.markdown('<div style="background-color: #F2CEEF; padding: 5px; text-align: center; border-radius: 3px;">Dry</div>', unsafe_allow_html=True)
            with legend_cols[4]:
                st.markdown('<div style="background-color: #00B0F0; padding: 5px; text-align: center; border-radius: 3px;">Unload</div>', unsafe_allow_html=True)
            
            # Export options
            st.markdown("---")
            st.subheader("💾 Export Schedule")
            
            # Create formatted Excel export
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_schedule = BytesIO()
            with pd.ExcelWriter(output_schedule, engine='openpyxl') as writer:
                # Write schedule
                schedule_export = schedule_df.copy()
                schedule_export.to_excel(writer, sheet_name='Schedule', index=False)
                
                # Write summary
                schedule_summary = {
                    'Metric': [
                        'Start Date',
                        'Days Simulated',
                        'Cell Size (CY)',
                        'Number of Cells',
                        'Daily Incoming Volume (CY)',
                        'Daily Equipment Capacity (CY)',
                        'Total Soil Loaded (CY)',
                        'Total Soil Unloaded (CY)',
                        'Soil Waiting (End)',
                    ],
                    'Value': [
                        schedule_start.strftime('%Y-%m-%d'),
                        len(schedule_df),
                        selected_config['cell_volume_cy'],
                        int(selected_config['num_cells']),
                        daily_volume,
                        daily_equipment_capacity,
                        schedule_df['CumSoilIn'].iloc[-1],
                        schedule_df['CumSoilOut'].iloc[-1],
                        schedule_df['SoilWaiting'].iloc[-1],
                    ]
                }
                pd.DataFrame(schedule_summary).to_excel(writer, sheet_name='Summary', index=False)
                
                # Apply formatting to Schedule sheet
                from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                worksheet = writer.sheets['Schedule']
                
                # Define colors matching display
                excel_phase_colors = {
                    'Load': '#8ED973',
                    'Rip': '#83CCEB',
                    'Treat': '#FFC000',
                    'Dry': '#F2CEEF',
                    'Unload': '#00B0F0',
                    'ReadyToUnload': '#00B0F0',
                    'Empty': '#FFFFFF'
                }
                
                sunday_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                aptos_font = Font(name='Aptos Narrow', size=10)
                aptos_font_bold = Font(name='Aptos Narrow', size=10, bold=True)
                center_aligned = Alignment(horizontal='center', vertical='center')
                
                # Find columns
                phase_columns = []
                date_column_idx = None
                dayname_column_idx = None
                
                for col_idx, col_name in enumerate(schedule_df.columns, start=1):
                    if 'Phase' in col_name:
                        phase_columns.append((col_idx, col_name))
                    if col_name == 'Date':
                        date_column_idx = col_idx
                    if col_name == 'DayName':
                        dayname_column_idx = col_idx
                
                # Apply formatting
                for row_idx in range(1, len(schedule_df) + 2):
                    is_sunday = False
                    
                    if row_idx > 1:
                        if dayname_column_idx:
                            day_name_cell = worksheet.cell(row=row_idx, column=dayname_column_idx)
                            is_sunday = str(day_name_cell.value) == 'Sunday'
                    
                    for col_idx in range(1, len(schedule_df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        is_phase_column = any(col_idx == phase_col_idx for phase_col_idx, _ in phase_columns)
                        
                        cell.border = thin_border
                        cell.font = aptos_font_bold if row_idx == 1 else aptos_font
                        
                        if col_idx == date_column_idx and row_idx > 1:
                            cell.number_format = 'M/D/YYYY'
                        
                        if is_sunday and not is_phase_column and row_idx > 1 and col_idx != date_column_idx:
                            cell.fill = sunday_fill
                        
                        if row_idx > 1:
                            for phase_col_idx, phase_col_name in phase_columns:
                                if col_idx == phase_col_idx:
                                    cell.alignment = center_aligned
                                    phase_value = str(cell.value) if cell.value else ''
                                    
                                    for phase_name, color in excel_phase_colors.items():
                                        if phase_name in phase_value:
                                            hex_color = color.lstrip('#')
                                            openpyxl_color = 'FF' + hex_color
                                            cell.fill = PatternFill(start_color=openpyxl_color,
                                                                  end_color=openpyxl_color,
                                                                  fill_type='solid')
                                            break
                
                # Auto-adjust column widths
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
            
            output_schedule.seek(0)
            
            schedule_filename = f"treatment_schedule_{timestamp}.xlsx"
            st.download_button(
                label="📥 Download Formatted Schedule (Excel)",
                data=output_schedule,
                file_name=schedule_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_schedule"
            )
    
    else:
        # Initial state - show instructions
        st.info("👈 Configure your facility parameters in the sidebar and click **Optimize Configuration**")
        
        st.markdown("""
        ### How This Tool Works
        
        This optimizer helps you design a continuous-flow soil remediation facility by:
        
        1. **Analyzing Your Requirements**: Input your daily incoming soil volume and operational parameters
        2. **Finding Viable Configurations**: Only shows configurations that can sustain your planned volume
        3. **Simulating Operations**: Generates a detailed day-by-day schedule with color-coded phases
        4. **You Choose**: Select the configuration that best fits your needs
        
        ### Key Features
        
        - **Smart Filtering**: Only shows configurations that work for your volume
        - **Visual Schedule**: Color-coded treatment phases matching Excel export
        - **Continuous Operation Check**: Ensures you never turn away work
        - **Weekend Scheduling**: Customizable work schedules by phase
        - **Excel Export**: Formatted schedule with colored cells
        
        ### Phase Colors
        
        - 🟢 **Load** (Green): Soil being loaded into cell
        - 🔵 **Rip** (Light Blue): Ripping/mixing phase
        - 🟡 **Treat** (Gold): Chemical treatment phase
        - 🩷 **Dry** (Pink): Drying phase
        - 🔷 **Unload** (Blue): Soil being removed from cell
        
        ### Workflow
        
        1. Set your daily incoming volume and equipment capacity
        2. Configure treatment phase durations
        3. Define weekend working schedules
        4. Click **Optimize Configuration**
        5. Select your preferred cell configuration
        6. Review the color-coded schedule
        7. Download the formatted Excel schedule
        """)


if __name__ == "__main__":
    main()
